import csv
import io
from datetime import timedelta

from django.contrib.auth.decorators import login_required
from django.db.models import Avg, Count, F, Q
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone

from apps.permissions.decorators import require_global_permission
from apps.permissions.services import PermissionService
from apps.accounts.models import User, UserRole
from apps.designs.models import DesignRequest, DesignStatus
from apps.projects.models import Project
from apps.reports.audit_report import format_audit_due_value, get_audit_report_for_design_number


def _designer_performance_data(designer_id=None, project_id=None):
    designers = PermissionService.get_design_team_members()
    if designer_id:
        designers = designers.filter(pk=designer_id)
    rows = []
    for d in designers:
        qs = DesignRequest.objects.filter(assigned_designer=d)
        if project_id:
            qs = qs.filter(project_id=project_id)
        assigned = qs.count()
        completed = qs.filter(status=DesignStatus.COMPLETED).count()
        on_time = qs.filter(status=DesignStatus.COMPLETED, completion_date__lte=F('due_date')).count()
        corrections = qs.filter(revision_count__gt=0).count()
        completed_qs = qs.filter(status=DesignStatus.COMPLETED, completion_date__isnull=False)
        day_total, day_count = 0, 0
        for dr in completed_qs[:100]:
            if dr.completion_date and dr.created_at:
                day_total += (dr.completion_date - dr.created_at).total_seconds() / 86400
                day_count += 1
        avg_days_val = round(day_total / day_count, 1) if day_count else 0
        rate = round((completed / assigned * 100) if assigned else 0, 1)
        on_time_rate = round((on_time / completed * 100) if completed else 0, 1)
        score = min(100, round(rate * 0.5 + on_time_rate * 0.5))
        rows.append({
            'designer': d.get_full_name() or d.username,
            'assigned': assigned,
            'completed': completed,
            'on_time': on_time,
            'completion_rate': rate,
            'on_time_rate': on_time_rate,
            'avg_days': avg_days_val,
            'corrections': corrections,
            'score': score,
        })
    return rows


def _deadline_compliance_data():
    designs = DesignRequest.objects.exclude(deadline_due__isnull=True)
    total = designs.count()
    breached = designs.filter(deadline_status='red').count()
    return {
        'total': total,
        'green': designs.filter(deadline_status='green').count(),
        'yellow': designs.filter(deadline_status='yellow').count(),
        'red': breached,
        'compliance_rate': round(((total - breached) / total * 100) if total else 100, 1),
    }


def _export_query(request, *keys):
    parts = []
    for key in keys:
        value = request.GET.get(key)
        if value:
            parts.append(f'{key}={value}')
    return '&'.join(parts)


def _append_query(url, query):
    if not query:
        return url
    separator = '&' if '?' in url else '?'
    return f'{url}{separator}{query}'


def _report_export_urls(request, report_type, *filter_keys):
    from django.urls import reverse
    query = _export_query(request, *filter_keys)
    return {
        'csv': _append_query(reverse('reports:export_csv', args=[report_type]), query),
        'excel': _append_query(reverse('reports:export_excel', args=[report_type]), query),
        'pdf': _append_query(reverse('reports:export_pdf', args=[report_type]), query),
    }


def _format_audit_timestamp(ts):
    if not ts:
        return ''
    return timezone.localtime(ts).strftime('%Y-%m-%d %H:%M')


AUDIT_TABLE_HEADERS = [
    'Stage', 'Actor', 'Role', 'Due Label', 'Due At', 'SLA Due At', 'Completed At',
    'Duration (days)', 'On Time Status', 'Status Note', 'Notes',
    'Delayed', 'Delayed By', 'Delay Days', 'Delay Type', 'Delay Note',
]


def _audit_status_display(row):
    status = row.get('on_time_status', 'n/a')
    mapping = {
        'late': 'LATE',
        'on_time': 'On Time',
        'due_set': 'Due Set',
        'n/a': 'N/A',
    }
    return mapping.get(status, status)


def _audit_row_values(row):
    stage = row['stage']
    if row.get('is_delayed'):
        stage = f'[LATE] {stage}'
    return [
        stage,
        row['actor'],
        row['role'],
        row.get('due_label', ''),
        format_audit_due_value(row.get('due_at')),
        format_audit_due_value(row.get('sla_due_at')),
        _format_audit_timestamp(row['timestamp']),
        row.get('duration_days', ''),
        _audit_status_display(row),
        row.get('status_note', ''),
        row.get('notes', ''),
        'Yes' if row.get('is_delayed') else 'No',
        row.get('delayed_by', ''),
        row.get('delay_days', ''),
        row.get('delay_type', ''),
        row.get('delay_note', ''),
    ]


def _write_audit_report_header_rows(writer, report):
    writer.writerow(['Design Number', report['design_number']])
    writer.writerow(['Project', f"{report['project_code']} — {report['project']}"])
    writer.writerow(['Requested By', report['requested_by']])
    writer.writerow([
        'Target Completion Date',
        report['target_completion_date'].isoformat() if report['target_completion_date'] else '',
    ])
    key_dates = report.get('key_dates') or {}
    writer.writerow([
        'Engineer Due',
        format_audit_due_value(key_dates.get('engineer_due')),
    ])
    writer.writerow([
        'Designer Due (HOD)',
        format_audit_due_value(key_dates.get('designer_due')),
    ])
    writer.writerow([
        'Verification Due',
        format_audit_due_value(key_dates.get('verification_due')),
    ])
    writer.writerow([
        'Compliance Due',
        format_audit_due_value(key_dates.get('compliance_due')),
    ])
    writer.writerow(['Status', report['status']])
    if report.get('delay_summary'):
        summary = report['delay_summary']
        writer.writerow(['Primary Delay Source', summary.get('primary_source', '')])
        writer.writerow(['Primary Delay Days', summary.get('primary_days', '')])


def _style_audit_excel_row(ws, row_idx, on_time_status):
    try:
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return
    fills = {
        'late': PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid'),
        'on_time': PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid'),
        'due_set': PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid'),
    }
    fill = fills.get(on_time_status)
    status_font = Font(color='B91C1C', bold=True) if on_time_status == 'late' else None
    for col in range(1, len(AUDIT_TABLE_HEADERS) + 1):
        cell = ws.cell(row=row_idx, column=col)
        if fill:
            cell.fill = fill
        if status_font and col == 9:
            cell.font = status_font


def _write_audit_excel_report(ws, report):
    rows = [
        ['Design Number', report['design_number']],
        ['Project', f"{report['project_code']} — {report['project']}"],
        ['Requested By', report['requested_by']],
        [
            'Target Completion Date',
            report['target_completion_date'].isoformat() if report['target_completion_date'] else '',
        ],
    ]
    key_dates = report.get('key_dates') or {}
    rows.extend([
        ['Engineer Due', format_audit_due_value(key_dates.get('engineer_due'))],
        ['Designer Due (HOD)', format_audit_due_value(key_dates.get('designer_due'))],
        ['Verification Due', format_audit_due_value(key_dates.get('verification_due'))],
        ['Compliance Due', format_audit_due_value(key_dates.get('compliance_due'))],
        ['Status', report['status']],
    ])
    if report.get('delay_summary'):
        summary = report['delay_summary']
        rows.append(['Primary Delay Source', summary.get('primary_source', '')])
        rows.append(['Primary Delay Days', summary.get('primary_days', '')])
    for row in rows:
        ws.append(row)
    ws.append([])
    header_row_idx = ws.max_row
    ws.append(AUDIT_TABLE_HEADERS)
    try:
        from openpyxl.styles import Font, PatternFill
        header_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
        header_font = Font(bold=True)
        for col in range(1, len(AUDIT_TABLE_HEADERS) + 1):
            cell = ws.cell(row=header_row_idx, column=col)
            cell.fill = header_fill
            cell.font = header_font
        ws.freeze_panes = ws.cell(row=header_row_idx + 1, column=1)
    except ImportError:
        pass
    for row in report['rows']:
        ws.append(_audit_row_values(row))
        _style_audit_excel_row(ws, ws.max_row, row.get('on_time_status', 'n/a'))


def _write_audit_report_rows(writer, report):
    _write_audit_report_header_rows(writer, report)
    writer.writerow([])
    writer.writerow(AUDIT_TABLE_HEADERS)
    for row in report['rows']:
        writer.writerow(_audit_row_values(row))


def _write_csv_rows(writer, report_type, designer_filter=None, project_filter=None, design_filter=None):
    if report_type == 'designer_performance':
        writer.writerow([
            'Designer', 'Assigned', 'Completed', 'On-Time', 'Completion Rate %',
            'On-Time Rate %', 'Avg Days', 'Corrections', 'Score',
        ])
        for row in _designer_performance_data(designer_filter, project_filter):
            writer.writerow([
                row['designer'], row['assigned'], row['completed'], row['on_time'],
                row['completion_rate'], row['on_time_rate'], row['avg_days'],
                row['corrections'], row['score'],
            ])
    elif report_type == 'project_progress':
        writer.writerow([
            'Project Code', 'Client Name', 'Status', 'Total Designs',
            'Completed', 'Running', 'Progress %', 'Health Score',
        ])
        for p in Project.objects.all():
            total = p.total_design_requests
            pct = round(p.completed_designs / total * 100) if total else 0
            writer.writerow([
                p.code, p.client_name, p.get_status_display(), total,
                p.completed_designs, p.running_designs, pct, p.health_score,
            ])
    elif report_type == 'deadline_compliance':
        writer.writerow(['Design Number', 'Project', 'Deadline Status', 'Due Date', 'Status'])
        for d in DesignRequest.objects.exclude(deadline_due__isnull=True).select_related('project'):
            writer.writerow([
                d.design_number, d.project.code, d.deadline_status,
                d.deadline_due, d.get_status_display(),
            ])
    elif report_type == 'delay_analysis':
        writer.writerow(['Design Number', 'Delay Source', 'Delay Days', 'Status'])
        for d in DesignRequest.objects.exclude(delay_source=''):
            writer.writerow([
                d.design_number, d.delay_source, d.delay_duration_days, d.get_status_display(),
            ])
    elif report_type == 'verification':
        writer.writerow(['Design Number', 'Verifier', 'Status', 'Revisions'])
        for d in DesignRequest.objects.filter(verified_by__isnull=False).select_related('verified_by'):
            writer.writerow([
                d.design_number,
                d.verified_by.get_full_name() if d.verified_by else '',
                d.get_status_display(), d.revision_count,
            ])
    elif report_type == 'design_workflow_audit':
        report = get_audit_report_for_design_number(design_filter)
        if not report:
            writer.writerow(['Error', f'Design not found: {design_filter or ""}'])
            return
        _write_audit_report_rows(writer, report)
    else:
        writer.writerow(['Metric', 'Value'])
        deadline_stats = _deadline_compliance_data()
        writer.writerow(['Total Projects', Project.objects.count()])
        writer.writerow(['Total Designs', DesignRequest.objects.count()])
        writer.writerow(['Pending Designs', DesignRequest.objects.exclude(
            status__in=[DesignStatus.COMPLETED, DesignStatus.CANCELLED],
        ).count()])
        writer.writerow(['Deadline Compliance %', deadline_stats['compliance_rate']])


def _write_excel_rows(ws, report_type, designer_filter=None, project_filter=None, design_filter=None):
    if report_type == 'designer_performance':
        ws.append([
            'Designer', 'Assigned', 'Completed', 'On-Time', 'Completion Rate %',
            'On-Time Rate %', 'Avg Days', 'Corrections', 'Score',
        ])
        for row in _designer_performance_data(designer_filter, project_filter):
            ws.append([
                row['designer'], row['assigned'], row['completed'], row['on_time'],
                row['completion_rate'], row['on_time_rate'], row['avg_days'],
                row['corrections'], row['score'],
            ])
    elif report_type == 'project_progress':
        ws.append([
            'Project Code', 'Client Name', 'Status', 'Total Designs',
            'Completed', 'Running', 'Progress %', 'Health Score',
        ])
        for p in Project.objects.all():
            total = p.total_design_requests
            pct = round(p.completed_designs / total * 100) if total else 0
            ws.append([
                p.code, p.client_name, p.get_status_display(), total,
                p.completed_designs, p.running_designs, pct, p.health_score,
            ])
    elif report_type == 'deadline_compliance':
        ws.append(['Design Number', 'Project', 'Deadline Status', 'Due Date', 'Status'])
        for d in DesignRequest.objects.exclude(deadline_due__isnull=True).select_related('project'):
            ws.append([
                d.design_number, d.project.code, d.deadline_status,
                d.deadline_due, d.get_status_display(),
            ])
    elif report_type == 'delay_analysis':
        ws.append(['Design Number', 'Delay Source', 'Delay Days', 'Status'])
        for d in DesignRequest.objects.exclude(delay_source=''):
            ws.append([
                d.design_number, d.delay_source, d.delay_duration_days, d.get_status_display(),
            ])
    elif report_type == 'verification':
        ws.append(['Design Number', 'Verifier', 'Status', 'Revisions'])
        for d in DesignRequest.objects.filter(verified_by__isnull=False).select_related('verified_by'):
            ws.append([
                d.design_number,
                d.verified_by.get_full_name() if d.verified_by else '',
                d.get_status_display(), d.revision_count,
            ])
    elif report_type == 'design_workflow_audit':
        report = get_audit_report_for_design_number(design_filter)
        if not report:
            ws.append(['Error', f'Design not found: {design_filter or ""}'])
            return
        _write_audit_excel_report(ws, report)
    else:
        deadline_stats = _deadline_compliance_data()
        ws.append(['Metric', 'Value'])
        ws.append(['Total Projects', Project.objects.count()])
        ws.append(['Total Designs', DesignRequest.objects.count()])
        ws.append(['Pending Designs', DesignRequest.objects.exclude(
            status__in=[DesignStatus.COMPLETED, DesignStatus.CANCELLED],
        ).count()])
        ws.append(['Deadline Compliance %', deadline_stats['compliance_rate']])


def _write_pdf_lines(p, report_type, designer_filter=None, project_filter=None, design_filter=None):
    p.setFont('Helvetica-Bold', 14)
    p.drawString(50, 800, f'Genesis Design - {report_type.replace("_", " ").title()}')
    p.setFont('Helvetica', 10)
    y = 770

    def next_line(text):
        nonlocal y
        p.drawString(50, y, text[:110])
        y -= 15
        if y < 50:
            p.showPage()
            p.setFont('Helvetica', 10)
            y = 800

    if report_type == 'designer_performance':
        for row in _designer_performance_data(designer_filter, project_filter):
            next_line(
                f"{row['designer']}: {row['completed']}/{row['assigned']} "
                f"({row['completion_rate']}%) · Score {row['score']}"
            )
    elif report_type == 'project_progress':
        for proj in Project.objects.all()[:80]:
            total = proj.total_design_requests
            pct = round(proj.completed_designs / total * 100) if total else 0
            next_line(f"{proj.code}: {proj.completed_designs}/{total} complete ({pct}%)")
    elif report_type == 'deadline_compliance':
        stats = _deadline_compliance_data()
        for line in [
            f"Compliance rate: {stats['compliance_rate']}%",
            f"On track: {stats['green']}",
            f"Warning: {stats['yellow']}",
            f"Missed: {stats['red']}",
            f"Total tracked: {stats['total']}",
        ]:
            next_line(line)
    elif report_type == 'delay_analysis':
        for d in DesignRequest.objects.exclude(delay_source='')[:80]:
            next_line(
                f"{d.design_number}: {d.delay_source} ({d.delay_duration_days} days)"
            )
    elif report_type == 'design_workflow_audit':
        report = get_audit_report_for_design_number(design_filter)
        if not report:
            next_line(f'Design not found: {design_filter or ""}')
            return
        next_line(f"Design: {report['design_number']} · {report['project_code']}")
        next_line(f"Requested by: {report['requested_by']}")
        key_dates = report.get('key_dates') or {}
        if key_dates.get('requester_target'):
            next_line(f"Requester target: {key_dates['requester_target'].isoformat()}")
        if key_dates.get('engineer_due'):
            next_line(f"Engineer due: {format_audit_due_value(key_dates['engineer_due'])}")
        if key_dates.get('designer_due'):
            next_line(f"Designer due (HOD): {format_audit_due_value(key_dates['designer_due'])}")
        if key_dates.get('verification_due'):
            next_line(f"Verification due: {format_audit_due_value(key_dates['verification_due'])}")
        if key_dates.get('compliance_due'):
            next_line(f"Compliance due: {format_audit_due_value(key_dates['compliance_due'])}")
        if report.get('delay_summary'):
            summary = report['delay_summary']
            next_line(
                f"Primary delay: {summary.get('primary_source', '')} "
                f"({summary.get('primary_days', '')} days)"
            )
        next_line('')
        try:
            from reportlab.lib import colors
        except ImportError:
            colors = None
        for row in report['rows']:
            status = row.get('on_time_status', 'n/a')
            due_part = ''
            if row.get('due_label'):
                due_part = f" · Due: {row['due_label']}"
                if row.get('due_at'):
                    due_part += f" {format_audit_due_value(row['due_at'])}"
            delay = ''
            suffix = ''
            if row.get('is_delayed') or status == 'late':
                delay = f" · DELAY: {row.get('delayed_by', '')} ({row.get('delay_note', '')})"
                prefix = '** LATE ** '
            elif status == 'on_time':
                prefix = ''
                suffix = ' [OK]'
            elif status == 'due_set':
                prefix = ''
                suffix = ' [Due Set]'
            else:
                prefix = ''
            line = (
                f"{prefix}{_format_audit_timestamp(row['timestamp'])} · {row['stage']} · "
                f"{row['actor']} ({row['role']}){due_part}{delay}{suffix}"
            )
            if colors and (row.get('is_delayed') or status == 'late'):
                p.setFillColor(colors.HexColor('#B91C1C'))
                next_line(line)
                p.setFillColor(colors.black)
            elif colors and status == 'on_time':
                p.setFillColor(colors.HexColor('#15803D'))
                next_line(line)
                p.setFillColor(colors.black)
            else:
                next_line(line)
    else:
        deadline_stats = _deadline_compliance_data()
        for line in [
            f'Total Projects: {Project.objects.count()}',
            f'Total Designs: {DesignRequest.objects.count()}',
            f'Deadline Compliance: {deadline_stats["compliance_rate"]}%',
            f'Missed Deadlines: {deadline_stats["red"]}',
        ]:
            next_line(line)


@login_required
@require_global_permission('NAV_PERM_REPORTS')
def reports_index(request):
    tab = request.GET.get('tab', 'performance')
    designer_filter = request.GET.get('designer')
    project_filter = request.GET.get('project')
    design_filter = (request.GET.get('design') or '').strip()
    audit_report = get_audit_report_for_design_number(design_filter) if design_filter else None
    delay_data = DesignRequest.objects.exclude(delay_source='').values(
        'design_number', 'delay_source', 'delay_duration_days', 'status'
    )[:50]
    deadline_data = _deadline_compliance_data()
    project_rows = []
    for p in Project.objects.all():
        project_rows.append({
            'code': p.code,
            'name': p.client_name,
            'status': p.get_status_display(),
            'total': p.total_design_requests,
            'completed': p.completed_designs,
            'running': p.running_designs,
            'health': p.health_score,
            'pct': round(p.completed_designs / p.total_design_requests * 100) if p.total_design_requests else 0,
        })

    performance_exports = _report_export_urls(
        request, 'designer_performance', 'designer', 'project',
    )
    project_exports = _report_export_urls(request, 'project_progress')
    deadline_exports = _report_export_urls(request, 'deadline_compliance')
    delay_exports = _report_export_urls(request, 'delay_analysis')
    summary_exports = _report_export_urls(request, 'management_summary')
    audit_exports = _report_export_urls(request, 'design_workflow_audit', 'design') if audit_report else None

    report_catalog = [
        {
            'title': 'Designer Performance',
            'description': 'Completion rates, on-time delivery, and quality scores by designer.',
            'icon': 'users',
            'icon_bg': 'bg-blue-50',
            'icon_color': 'text-blue-600',
            'exports': performance_exports,
        },
        {
            'title': 'Project Progress',
            'description': 'Design volume, completion progress, and health scores per project.',
            'icon': 'folder-kanban',
            'icon_bg': 'bg-indigo-50',
            'icon_color': 'text-indigo-600',
            'exports': project_exports,
        },
        {
            'title': 'Deadline Compliance',
            'description': 'On-track, warning, and breached deadline status across designs.',
            'icon': 'calendar-clock',
            'icon_bg': 'bg-amber-50',
            'icon_color': 'text-amber-600',
            'exports': deadline_exports,
        },
        {
            'title': 'Delay Analysis',
            'description': 'Delay attribution by source with duration and current status.',
            'icon': 'timer',
            'icon_bg': 'bg-red-50',
            'icon_color': 'text-red-600',
            'exports': delay_exports,
        },
        {
            'title': 'Management Summary',
            'description': 'Executive snapshot of projects, designs, and compliance metrics.',
            'icon': 'pie-chart',
            'icon_bg': 'bg-green-50',
            'icon_color': 'text-green-600',
            'exports': {
                'csv': summary_exports['csv'],
                'excel': summary_exports['excel'],
                'pdf': summary_exports['pdf'],
            },
        },
        {
            'title': 'Design Workflow Audit',
            'description': 'Per-request stage timeline with actors, timestamps, and delay attribution.',
            'icon': 'clipboard-list',
            'icon_bg': 'bg-violet-50',
            'icon_color': 'text-violet-600',
            'exports': audit_exports or {
                'csv': '',
                'excel': '',
                'pdf': '',
            },
            'requires_design': True,
        },
    ]

    return render(request, 'reports/index.html', {
        'tab': tab,
        'designer_data': _designer_performance_data(designer_filter, project_filter),
        'deadline_data': deadline_data,
        'delay_data': delay_data,
        'project_data': project_rows,
        'designers': PermissionService.get_design_team_members(),
        'projects': Project.objects.all(),
        'designer_filter': designer_filter,
        'project_filter': project_filter,
        'total_projects': Project.objects.count(),
        'total_designs': DesignRequest.objects.count(),
        'pending_designs': DesignRequest.objects.exclude(
            status__in=[DesignStatus.COMPLETED, DesignStatus.CANCELLED]
        ).count(),
        'performance_exports': performance_exports,
        'project_exports': project_exports,
        'deadline_exports': deadline_exports,
        'delay_exports': delay_exports,
        'summary_exports': summary_exports,
        'audit_exports': audit_exports,
        'audit_report': audit_report,
        'design_filter': design_filter,
        'report_catalog': report_catalog,
    })


@login_required
@require_global_permission('NAV_PERM_REPORTS')
def export_csv(request, report_type):
    designer_filter = request.GET.get('designer')
    project_filter = request.GET.get('project')
    design_filter = request.GET.get('design')
    response = HttpResponse(content_type='text/csv')
    filename = report_type
    if report_type == 'design_workflow_audit' and design_filter:
        filename = f'design_workflow_audit_{design_filter}'
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    writer = csv.writer(response)
    _write_csv_rows(writer, report_type, designer_filter, project_filter, design_filter)
    return response


@login_required
@require_global_permission('NAV_PERM_REPORTS')
def export_excel(request, report_type):
    designer_filter = request.GET.get('designer')
    project_filter = request.GET.get('project')
    design_filter = request.GET.get('design')
    try:
        import openpyxl
    except ImportError:
        return export_csv(request, report_type)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_type[:31]
    _write_excel_rows(ws, report_type, designer_filter, project_filter, design_filter)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = report_type
    if report_type == 'design_workflow_audit' and design_filter:
        filename = f'design_workflow_audit_{design_filter}'
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)
    return response


@login_required
@require_global_permission('NAV_PERM_REPORTS')
def export_pdf(request, report_type):
    designer_filter = request.GET.get('designer')
    project_filter = request.GET.get('project')
    design_filter = request.GET.get('design')
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
    except ImportError:
        return export_csv(request, report_type)

    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    _write_pdf_lines(p, report_type, designer_filter, project_filter, design_filter)
    p.showPage()
    p.save()
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    filename = report_type
    if report_type == 'design_workflow_audit' and design_filter:
        filename = f'design_workflow_audit_{design_filter}'
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    return response
